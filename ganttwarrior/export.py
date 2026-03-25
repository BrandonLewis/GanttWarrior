"""Export functionality: PDF, Excel, and terminal print."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import Optional

from rich.console import Console
from rich.table import Table
from rich.text import Text

from .models import Project, Task, TaskColor, TaskStatus
from .views.gantt import COLOR_MAP, STATUS_SYMBOLS


def _build_task_table(project: Project, title: Optional[str] = None) -> Table:
    """Build a Rich table of project tasks."""
    table = Table(title=title or project.name, show_lines=True)
    table.add_column("WBS", style="bold", width=8)
    table.add_column("Task", width=30)
    table.add_column("Status", width=12)
    table.add_column("Start", width=12)
    table.add_column("End", width=12)
    table.add_column("Days", width=6, justify="right")
    table.add_column("Progress", width=10, justify="right")
    table.add_column("Assigned", width=15)
    table.add_column("Critical", width=8, justify="center")
    table.add_column("Deps", width=15)

    for task in project.sorted_tasks():
        color = COLOR_MAP.get(task.color, "white")
        status_sym = STATUS_SYMBOLS.get(task.status, "?")

        deps = ", ".join(d.predecessor_id[:6] for d in task.dependencies) if task.dependencies else ""
        progress_str = f"{int(task.progress * 100)}%"

        table.add_row(
            task.wbs,
            Text(task.name, style=color),
            f"{status_sym} {task.status.value}",
            task.start_date.isoformat() if task.start_date else "-",
            task.end_date.isoformat() if task.end_date else "-",
            str(task.duration_days),
            progress_str,
            task.assigned_to or "-",
            "⚠ YES" if task.is_critical else "",
            deps or "-",
        )

    return table


def _build_gantt_text(project: Project, day_width: int = 2) -> str:
    """Build a text-based Gantt chart for export."""
    tasks = project.sorted_tasks()
    if not tasks:
        return "No tasks in project."

    starts = [t.start_date for t in tasks if t.start_date]
    ends = [t.end_date for t in tasks if t.end_date]
    if not starts:
        return "No tasks with dates."

    chart_start = min(starts) - timedelta(days=1)
    chart_end = max(ends) + timedelta(days=1) if ends else min(starts) + timedelta(days=30)
    total_days = (chart_end - chart_start).days + 1
    label_width = 35

    lines = []

    # Header
    header = " " * label_width + "│"
    for i in range(total_days):
        d = chart_start + timedelta(days=i)
        if d.day == 1 or i == 0:
            month_str = d.strftime("%b")
            header += month_str[:day_width].ljust(day_width)
        else:
            header += str(d.day % 10) * day_width if day_width == 1 else str(d.day).rjust(day_width)
    lines.append(header)
    lines.append("─" * label_width + "┼" + "─" * (total_days * day_width))

    # Task rows
    for task in tasks:
        indent = "  " * max(task.wbs_level - 1, 0)
        status_sym = STATUS_SYMBOLS.get(task.status, "?")
        label = f"{indent}{status_sym} {task.wbs} {task.name}"
        if len(label) > label_width - 2:
            label = label[:label_width - 5] + "..."
        line = label.ljust(label_width) + "│"

        if task.start_date and task.end_date:
            start_off = max((task.start_date - chart_start).days, 0)
            end_off = min((task.end_date - chart_start).days + 1, total_days)
            bar_len = max(end_off - start_off, 0)
            line += " " * (start_off * day_width)
            if task.is_milestone:
                line += "◆"
            else:
                filled = int(bar_len * day_width * task.progress)
                remaining = bar_len * day_width - filled
                line += "█" * filled + "░" * remaining
            after = total_days * day_width - end_off * day_width
            if after > 0:
                line += " " * after
        else:
            line += " " * (total_days * day_width)

        if task.is_critical:
            line += " ◄ CRITICAL"
        lines.append(line)

    return "\n".join(lines)


def print_project(project: Project) -> None:
    """Print the project to terminal using Rich."""
    console = Console()
    console.print()
    console.print(_build_task_table(project))
    console.print()
    console.print(_build_gantt_text(project))
    console.print()

    # Print summary
    total = len(project.tasks)
    completed = sum(1 for t in project.tasks if t.status == TaskStatus.COMPLETED)
    in_progress = sum(1 for t in project.tasks if t.status == TaskStatus.IN_PROGRESS)
    blocked = sum(1 for t in project.tasks if t.status == TaskStatus.BLOCKED)
    critical = sum(1 for t in project.tasks if t.is_critical)

    console.print(f"\n[bold]Project Summary:[/bold]")
    console.print(f"  Total tasks: {total}")
    console.print(f"  Completed: [green]{completed}[/green]")
    console.print(f"  In Progress: [blue]{in_progress}[/blue]")
    console.print(f"  Blocked: [red]{blocked}[/red]")
    console.print(f"  Critical Path: [bold red]{critical} tasks[/bold red]")


def export_pdf(project: Project, path: str) -> str:
    """Export project to PDF using fpdf2."""
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A3")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, project.name, ln=True, align="C")
    pdf.ln(5)

    # Task table
    pdf.set_font("Helvetica", "B", 8)
    col_widths = [15, 60, 25, 25, 25, 12, 15, 30, 15, 30]
    headers = ["WBS", "Task", "Status", "Start", "End", "Days", "Progress", "Assigned", "Critical", "Dependencies"]

    for w, h in zip(col_widths, headers):
        pdf.cell(w, 7, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for task in project.sorted_tasks():
        status_str = task.status.value
        progress_str = f"{int(task.progress * 100)}%"
        deps = ", ".join(d.predecessor_id[:6] for d in task.dependencies) if task.dependencies else "-"
        indent = "  " * max(task.wbs_level - 1, 0)

        row = [
            task.wbs,
            indent + task.name,
            status_str,
            task.start_date.isoformat() if task.start_date else "-",
            task.end_date.isoformat() if task.end_date else "-",
            str(task.duration_days),
            progress_str,
            task.assigned_to or "-",
            "YES" if task.is_critical else "",
            deps,
        ]

        # Highlight critical tasks
        if task.is_critical:
            pdf.set_fill_color(255, 230, 230)
            fill = True
        else:
            fill = False

        for w, val in zip(col_widths, row):
            pdf.cell(w, 6, val[:int(w / 1.5)], border=1, fill=fill)
        pdf.ln()

    # Gantt chart section
    pdf.ln(10)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 10, "Gantt Chart", ln=True)

    tasks = project.sorted_tasks()
    if tasks:
        starts = [t.start_date for t in tasks if t.start_date]
        ends = [t.end_date for t in tasks if t.end_date]

        if starts and ends:
            chart_start = min(starts)
            chart_end = max(ends)
            total_days = (chart_end - chart_start).days + 1
            label_w = 70
            chart_w = min(300, total_days * 2)
            day_w = chart_w / max(total_days, 1)
            y_start = pdf.get_y() + 5

            pdf.set_font("Helvetica", "", 6)

            for i, task in enumerate(tasks):
                y = y_start + i * 5
                if y > pdf.h - 20:
                    pdf.add_page()
                    y_start = 20
                    y = y_start + i * 5

                # Task label
                pdf.set_xy(10, y)
                label = f"{task.wbs} {task.name}"
                pdf.cell(label_w, 4, label[:40], align="L")

                # Bar
                if task.start_date and task.end_date:
                    bar_start = label_w + 10 + (task.start_date - chart_start).days * day_w
                    bar_width = max((task.end_date - task.start_date).days + 1, 1) * day_w

                    # Color based on status
                    if task.is_critical:
                        pdf.set_fill_color(220, 50, 50)
                    elif task.status == TaskStatus.COMPLETED:
                        pdf.set_fill_color(50, 180, 50)
                    elif task.status == TaskStatus.IN_PROGRESS:
                        pdf.set_fill_color(50, 120, 220)
                    elif task.status == TaskStatus.BLOCKED:
                        pdf.set_fill_color(200, 100, 50)
                    else:
                        pdf.set_fill_color(100, 100, 180)

                    pdf.rect(bar_start, y, bar_width, 4, "F")

                    # Progress overlay
                    if task.progress > 0:
                        pdf.set_fill_color(0, 0, 0)
                        pdf.set_draw_color(0, 0, 0)
                        progress_w = bar_width * task.progress
                        pdf.rect(bar_start, y + 3.5, progress_w, 0.5, "F")

    # Summary
    pdf.ln(15)
    pdf.set_font("Helvetica", "B", 10)
    total = len(project.tasks)
    completed = sum(1 for t in project.tasks if t.status == TaskStatus.COMPLETED)
    critical = sum(1 for t in project.tasks if t.is_critical)
    pdf.cell(0, 6, f"Total: {total} tasks | Completed: {completed} | Critical path: {critical} tasks", ln=True)

    output = Path(path)
    pdf.output(str(output))
    return str(output)


def export_excel(project: Project, path: str) -> str:
    """Export project to Excel using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Task List Sheet ---
    ws = wb.active
    ws.title = "Tasks"

    headers = ["WBS", "Task Name", "Description", "Status", "Start Date", "End Date",
               "Duration (days)", "Progress %", "Assigned To", "Priority",
               "Critical Path", "Dependencies", "Color"]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    critical_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    completed_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
    blocked_fill = PatternFill(start_color="FFE8D0", end_color="FFE8D0", fill_type="solid")

    for row_idx, task in enumerate(project.sorted_tasks(), 2):
        deps = "; ".join(
            f"{d.predecessor_id}({d.dependency_type.value})"
            for d in task.dependencies
        ) if task.dependencies else ""

        values = [
            task.wbs,
            task.name,
            task.description,
            task.status.value,
            task.start_date,
            task.end_date,
            task.duration_days,
            task.progress * 100,
            task.assigned_to,
            task.priority,
            "Yes" if task.is_critical else "No",
            deps,
            task.color.value,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if task.is_critical:
                cell.fill = critical_fill
            elif task.status == TaskStatus.COMPLETED:
                cell.fill = completed_fill
            elif task.status == TaskStatus.BLOCKED:
                cell.fill = blocked_fill

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Gantt Data Sheet ---
    ws2 = wb.create_sheet("Gantt Data")
    ws2.cell(row=1, column=1, value="Task").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Start Offset (days)").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Duration (days)").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="Float (days)").font = Font(bold=True)

    tasks = project.sorted_tasks()
    if tasks:
        project_start = min((t.start_date for t in tasks if t.start_date), default=date.today())
        for row_idx, task in enumerate(tasks, 2):
            ws2.cell(row=row_idx, column=1, value=f"{task.wbs} {task.name}")
            offset = (task.start_date - project_start).days if task.start_date else 0
            ws2.cell(row=row_idx, column=2, value=offset)
            ws2.cell(row=row_idx, column=3, value=task.duration_days)
            ws2.cell(row=row_idx, column=4, value=task.total_float)

    # --- Kanban Sheet ---
    ws3 = wb.create_sheet("Kanban")
    statuses = ["Not Started", "Blocked", "In Progress", "Completed"]
    status_map = {
        "Not Started": TaskStatus.NOT_STARTED,
        "Blocked": TaskStatus.BLOCKED,
        "In Progress": TaskStatus.IN_PROGRESS,
        "Completed": TaskStatus.COMPLETED,
    }

    for col, status_name in enumerate(statuses, 1):
        cell = ws3.cell(row=1, column=col, value=status_name)
        cell.font = Font(bold=True, size=11)
        cell.fill = header_fill
        cell.font = header_font

        status = status_map[status_name]
        status_tasks = [t for t in project.sorted_tasks() if t.status == status]
        for row_idx, task in enumerate(status_tasks, 2):
            ws3.cell(row=row_idx, column=col, value=f"{task.wbs} {task.name}")

    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 35

    output = Path(path)
    wb.save(str(output))
    return str(output)
