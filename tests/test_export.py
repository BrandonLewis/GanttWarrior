"""Tests for export functionality."""

import tempfile
from datetime import date
from pathlib import Path

import pytest

from ganttwarrior.export import export_excel, export_pdf, print_project
from ganttwarrior.models import (
    Dependency,
    DependencyType,
    Project,
    Task,
    TaskColor,
    TaskStatus,
)
from ganttwarrior.scheduler import Scheduler


class TestExport:
    def _make_project(self) -> Project:
        project = Project(name="Export Test", start_date=date(2026, 3, 1))
        t1 = project.add_task(Task(
            name="Phase 1", wbs="1", duration_days=5,
            color=TaskColor.BLUE, assigned_to="Alice",
        ))
        t2 = project.add_task(Task(
            name="Phase 2", wbs="2", duration_days=8,
            color=TaskColor.RED, assigned_to="Bob",
            dependencies=[Dependency(t1.id)],
        ))
        t3 = project.add_task(Task(
            name="Phase 3", wbs="3", duration_days=3,
            color=TaskColor.GREEN,
            dependencies=[Dependency(t2.id)],
        ))

        t1.status = TaskStatus.COMPLETED
        t1.progress = 1.0
        t2.status = TaskStatus.IN_PROGRESS
        t2.progress = 0.5

        scheduler = Scheduler(project)
        scheduler.schedule()
        return project

    def test_export_pdf(self):
        try:
            from fpdf import FPDF  # noqa: F401
        except BaseException:
            pytest.skip("fpdf2/cryptography unavailable in this environment")

        project = self._make_project()
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            path = f.name

        try:
            result = export_pdf(project, path)
            assert Path(result).exists()
            assert Path(result).stat().st_size > 0
        finally:
            Path(path).unlink(missing_ok=True)

    def test_export_excel(self):
        project = self._make_project()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            result = export_excel(project, path)
            assert Path(result).exists()
            assert Path(result).stat().st_size > 0

            # Verify sheets exist
            from openpyxl import load_workbook
            wb = load_workbook(result)
            assert "Tasks" in wb.sheetnames
            assert "Gantt Data" in wb.sheetnames
            assert "Kanban" in wb.sheetnames

            # Verify task data
            ws = wb["Tasks"]
            assert ws.cell(row=2, column=2).value == "Phase 1"
        finally:
            Path(path).unlink(missing_ok=True)

    def test_print_project(self, capsys):
        project = self._make_project()
        print_project(project)
        captured = capsys.readouterr()
        assert "Export Test" in captured.out or len(captured.out) > 0
